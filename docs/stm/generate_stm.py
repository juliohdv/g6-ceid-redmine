"""Genera el documento de mapeo origen-destino (Source-to-Target Mapping) del modelo BI de Redmine.

Salida: redmine-stm.xlsx, una hoja por tabla del modelo dimensional.
Los tipos de dato de origen se tomaron de information_schema de redmine_db (MySQL 5.7).

Uso:
    python docs/stm/generate_stm.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- Constantes de origen -----------------------------------------------------
SRC = "Redmine"
SCH = "redmine_db"
DER = "Derived"

TARGET_COLS = [
    ("Column Name", 26),
    ("Display Name", 22),
    ("Description", 46),
    ("Datatype", 11),
    ("Size", 7),
    ("Precision", 10),
    ("Key?", 7),
    ("FK To", 26),
    ("NULL?", 7),
    ("Default Value", 13),
    ("Example Values", 20),
]

SOURCE_COLS = [
    ("Source System", 13),
    ("Source Schema", 13),
    ("Source Table", 20),
    ("Source Field Name", 20),
    ("Source Datatype", 14),
    ("Extraction/Transformation Rules", 54),
    ("Comments", 44),
]


def row(col, disp, desc, dtype="", size="", prec="", key="", fk="", nullable="N",
        default="", example="", src_sys="", src_schema="", src_table="", src_field="",
        src_dtype="", rules="", comments=""):
    return (col, disp, desc, dtype, size, prec, key, fk, nullable, default, example,
            src_sys, src_schema, src_table, src_field, src_dtype, rules, comments)


# =============================================================================
# TABLAS DE HECHOS
# =============================================================================

FACT_ISSUE = {
    "sheet": "FactIssue",
    "table_name": "FactIssue",
    "table_type": "Fact",
    "view_name": "vw_fact_issue",
    "display_name": "Issues",
    "description": ("FactIssue captura el ciclo de vida de las tareas al nivel de un issue "
                    "individual. Grano: una fila por issues.id."),
    "rows": [
        row("issue_key", "Issue Key", "Identificador del issue en el sistema transaccional",
            "int", key="PK", nullable="N", example="1, 2, 3",
            src_sys=SRC, src_schema=SCH, src_table="issues", src_field="id", src_dtype="int(11)",
            rules="Copia directa",
            comments="Dimension degenerada: el numero de issue es visible para el usuario final"),
        row("project_key", "Project Key", "Clave a la dimension Proyecto",
            "int", key="FK", fk="DimProject.project_key", nullable="N", example="1, 2, 3",
            src_sys=SRC, src_schema=SCH, src_table="issues", src_field="project_id", src_dtype="int(11)",
            rules="Copia directa",
            comments="Usar siempre issues.project_id para atribuir el issue, nunca versions.project_id (§5.3)"),
        row("tracker_key", "Tracker Key", "Clave a la dimension Tracker (Bug, Feature, Support)",
            "int", key="FK", fk="DimTracker.tracker_key", nullable="N", example="1, 2, 3",
            src_sys=SRC, src_schema=SCH, src_table="issues", src_field="tracker_id", src_dtype="int(11)",
            rules="Copia directa"),
        row("status_key", "Status Key", "Clave a la dimension Estado del issue",
            "int", key="FK", fk="DimIssueStatus.status_key", nullable="N", example="1, 2, 5",
            src_sys=SRC, src_schema=SCH, src_table="issues", src_field="status_id", src_dtype="int(11)",
            rules="Copia directa",
            comments="Estado vigente unicamente. La evolucion historica se modela en FactIssueHistory"),
        row("priority_key", "Priority Key", "Clave a la dimension Prioridad",
            "int", key="FK", fk="DimPriority.priority_key", nullable="N", example="1, 2, 3",
            src_sys=SRC, src_schema=SCH, src_table="issues", src_field="priority_id", src_dtype="int(11)",
            rules="Copia directa",
            comments="Actua como proxy de severidad mientras no exista el campo personalizado (§7.2)"),
        row("version_key", "Version Key", "Clave a la dimension Version (sprint o release nativo)",
            "int", key="FK", fk="DimVersion.version_key", nullable="Y", default="-1",
            example="1, 2, -1",
            src_sys=SRC, src_schema=SCH, src_table="issues", src_field="fixed_version_id", src_dtype="int(11)",
            rules="Copia directa. NULL se sustituye por -1",
            comments="NULL significa issue sin sprint asignado; se agrupa en el bucket 'Sin version' (§5.5)"),
        row("severity_key", "Severity Key", "Clave a la dimension Severidad (campo personalizado)",
            "int", key="FK", fk="DimSeverity.severity_key", nullable="Y", default="-1",
            example="1, 2, -1",
            src_sys=SRC, src_schema=SCH, src_table="custom_values", src_field="value", src_dtype="longtext",
            rules=("Lookup EAV: custom_values.customized_type = 'Issue' "
                   "AND custom_values.customized_id = issues.id "
                   "AND custom_fields.name = 'Severidad'"),
            comments="CONDICIONAL: el campo personalizado no existe hoy en el entorno. Verificar antes de cargar"),
        row("author_key", "Author Key", "Clave a la dimension Usuario que creo el issue",
            "int", key="FK", fk="DimUser.user_key", nullable="N", example="1, 5, 6",
            src_sys=SRC, src_schema=SCH, src_table="issues", src_field="author_id", src_dtype="int(11)",
            rules="Copia directa",
            comments="Role-playing dimension sobre DimUser; requiere rol inactivo o tabla duplicada en Power BI"),
        row("assignee_key", "Assignee Key", "Clave a la dimension Usuario responsable del issue",
            "int", key="FK", fk="DimUser.user_key", nullable="Y", default="-1",
            example="5, 6, -1",
            src_sys=SRC, src_schema=SCH, src_table="issues", src_field="assigned_to_id", src_dtype="int(11)",
            rules="Copia directa. NULL se sustituye por -1",
            comments="Role-playing dimension. NULL significa issue sin responsable asignado"),
        row("created_date_key", "Created Date Key", "Clave a la dimension Fecha para la fecha de creacion",
            "int", key="FK", fk="DimDate.date_key", nullable="N", example="20260701",
            src_sys=DER, src_table="issues", src_field="created_on", src_dtype="timestamp",
            rules="CONVERT(DATE(issues.created_on), formato YYYYMMDD)",
            comments="created_on se almacena en UTC del servidor; convertir a hora local si el informe lo exige (§1.3)"),
        row("due_date_key", "Due Date Key", "Clave a la dimension Fecha para la fecha comprometida",
            "int", key="FK", fk="DimDate.date_key", nullable="Y", default="-1",
            example="20260731, -1",
            src_sys=DER, src_table="issues", src_field="due_date", src_dtype="date",
            rules="CONVERT(issues.due_date, formato YYYYMMDD). NULL se sustituye por -1",
            comments="Sin due_date el KPI Schedule Variance no es calculable; excluir esos issues (§5.5)"),
        row("closed_date_key", "Closed Date Key", "Clave a la dimension Fecha para la fecha de cierre",
            "int", key="FK", fk="DimDate.date_key", nullable="Y", default="-1",
            example="20260712, -1",
            src_sys=DER, src_table="issues", src_field="closed_on", src_dtype="datetime",
            rules="CONVERT(DATE(issues.closed_on), formato YYYYMMDD). NULL se sustituye por -1",
            comments="Eje temporal del KPI Throughput"),
        row("issue_subject", "Issue Subject", "Titulo descriptivo del issue",
            "varchar", size="255", nullable="N", example="'Setup CI pipeline'",
            src_sys=SRC, src_schema=SCH, src_table="issues", src_field="subject", src_dtype="varchar(255)",
            rules="Copia directa",
            comments="Atributo degenerado; util para drill-through al detalle"),
        row("parent_issue_key", "Parent Issue Key", "Issue padre dentro de la jerarquia de subtareas",
            "int", nullable="Y", example="4, NULL",
            src_sys=SRC, src_schema=SCH, src_table="issues", src_field="parent_id", src_dtype="int(11)",
            rules="Copia directa",
            comments="Contar sin filtrar el nivel duplica metricas entre padre e hijos (§5.2)"),
        row("is_root_issue", "Is Root Issue", "Indica si el issue es raiz, es decir no es subtarea",
            "bit", nullable="N", default="1", example="0, 1",
            src_sys=DER, src_table="issues", src_field="parent_id", src_dtype="int(11)",
            rules="CASE WHEN issues.parent_id IS NULL THEN 1 ELSE 0 END",
            comments="Bandera de filtro obligatoria en Throughput y en el conteo de Burndown (§5.2)"),
        row("is_private", "Is Private", "Indica si el issue tiene visibilidad restringida",
            "bit", nullable="N", default="0", example="0, 1",
            src_sys=SRC, src_schema=SCH, src_table="issues", src_field="is_private", src_dtype="tinyint(1)",
            rules="Copia directa",
            comments="Evaluar seguridad a nivel de fila (RLS) en Power BI si varios equipos consumen el informe (§5.4)"),
        row("is_closed", "Is Closed", "Indica si el issue esta en un estado considerado cerrado",
            "bit", nullable="N", default="0", example="0, 1",
            src_sys=DER, src_table="issue_statuses", src_field="is_closed", src_dtype="tinyint(1)",
            rules="JOIN issue_statuses ON issue_statuses.id = issues.status_id, tomar is_closed",
            comments="Definicion canonica de tarea terminada para Throughput y Velocity (§5.1)"),
        row("estimated_hours", "Estimated Hours", "Horas estimadas de esfuerzo para el issue",
            "decimal", prec="9,2", nullable="Y", example="8.00, 16.00",
            src_sys=SRC, src_schema=SCH, src_table="issues", src_field="estimated_hours", src_dtype="float",
            rules="Copia directa con conversion a decimal(9,2)",
            comments="NULL frecuente. Excluir o imputar 0 segun politica y documentarlo en el informe (§5.5)"),
        row("done_ratio", "Done Ratio", "Porcentaje de avance declarado, de 0 a 100",
            "int", nullable="N", default="0", example="0, 40, 100",
            src_sys=SRC, src_schema=SCH, src_table="issues", src_field="done_ratio", src_dtype="int(11)",
            rules="Copia directa",
            comments="Puede diferir del cierre binario is_closed; son metricas independientes (§2.2)"),
        row("actual_hours", "Actual Hours", "Horas reales imputadas al issue",
            "decimal", prec="9,2", nullable="N", default="0", example="9.00, 13.00",
            src_sys=DER, src_table="time_entries", src_field="hours", src_dtype="float",
            rules="SUM(time_entries.hours) WHERE time_entries.issue_id = issues.id",
            comments="Excluye los time_entries con issue_id NULL, que se reportan a nivel proyecto (§6.4)"),
        row("hours_variance", "Hours Variance", "Desvio entre horas reales y horas estimadas",
            "decimal", prec="9,2", nullable="Y", example="1.50, -3.00",
            src_sys=DER,
            rules="actual_hours - estimated_hours",
            comments="NULL cuando no hay estimacion; presentar esos issues como 'Sin estimacion' (§6.4)"),
        row("hours_variance_pct", "Hours Variance Pct", "Desvio de horas expresado como porcentaje",
            "decimal", prec="9,4", nullable="Y", example="0.1875, -0.25",
            src_sys=DER,
            rules="DIVIDE(actual_hours - estimated_hours, estimated_hours)",
            comments="Usar DIVIDE para evitar division por cero cuando estimated_hours es 0 o NULL"),
        row("schedule_variance_days", "Schedule Variance Days",
            "Dias de desvio entre el cierre real y la fecha comprometida",
            "int", nullable="Y", example="1, -2",
            src_sys=DER, src_table="issues", src_field="closed_on, due_date", src_dtype="datetime, date",
            rules="DATEDIFF(issues.closed_on, issues.due_date)",
            comments="Solo issues cerrados con due_date informada. Positivo = retraso, negativo = adelanto (§6.1)"),
        row("remaining_hours", "Remaining Hours", "Horas pendientes segun el avance declarado",
            "decimal", prec="9,2", nullable="Y", example="9.60, 0.00",
            src_sys=DER,
            rules="estimated_hours * (1 - done_ratio / 100)",
            comments="Eje Y del burndown por snapshot actual, escenario A (§6.2)"),
        row("open_issue_count", "Open Issue Count", "Contador aditivo de issues abiertos",
            "int", nullable="N", default="0", example="0, 1",
            src_sys=DER,
            rules="CASE WHEN is_closed = 0 THEN 1 ELSE 0 END",
            comments="Medida aditiva; permite sumar en lugar de contar filas con filtro"),
    ],
}

FACT_TIME_ENTRY = {
    "sheet": "FactTimeEntry",
    "table_name": "FactTimeEntry",
    "table_type": "Fact",
    "view_name": "vw_fact_time_entry",
    "display_name": "Registros de tiempo",
    "description": ("FactTimeEntry captura el esfuerzo real registrado por los miembros del equipo. "
                    "Grano: una fila por time_entries.id."),
    "rows": [
        row("time_entry_key", "Time Entry Key", "Identificador del registro de tiempo en el origen",
            "int", key="PK", nullable="N", example="1, 2, 3",
            src_sys=SRC, src_schema=SCH, src_table="time_entries", src_field="id", src_dtype="int(11)",
            rules="Copia directa", comments="Dimension degenerada"),
        row("project_key", "Project Key", "Clave a la dimension Proyecto",
            "int", key="FK", fk="DimProject.project_key", nullable="N", example="1, 2",
            src_sys=SRC, src_schema=SCH, src_table="time_entries", src_field="project_id", src_dtype="int(11)",
            rules="Copia directa"),
        row("issue_key", "Issue Key", "Clave al hecho Issue sobre el que se imputo el tiempo",
            "int", key="FK", fk="FactIssue.issue_key", nullable="Y", default="-1",
            example="1, 4, -1",
            src_sys=SRC, src_schema=SCH, src_table="time_entries", src_field="issue_id", src_dtype="int(11)",
            rules="Copia directa. NULL se sustituye por -1",
            comments="NULL representa horas imputadas al proyecto sin tarea concreta; reportarlas aparte (§5.1)"),
        row("user_key", "User Key", "Clave a la dimension Usuario que registro las horas",
            "int", key="FK", fk="DimUser.user_key", nullable="N", example="5, 6, 7",
            src_sys=SRC, src_schema=SCH, src_table="time_entries", src_field="user_id", src_dtype="int(11)",
            rules="Copia directa",
            comments="time_entries.author_id puede diferir de user_id cuando un tercero carga las horas"),
        row("activity_key", "Activity Key", "Clave al tipo de actividad imputada",
            "int", key="FK", fk="DimActivity.activity_key", nullable="N", example="8, 9",
            src_sys=SRC, src_schema=SCH, src_table="time_entries", src_field="activity_id", src_dtype="int(11)",
            rules="Copia directa",
            comments=("BRECHA DEL MODELO: redmine-bi-model.md §3.3 declara ActivityKey pero §3.2 no define "
                      "DimActivity. Fuente sugerida: enumerations WHERE type = 'TimeEntryActivity'")),
        row("spent_on_date_key", "Spent On Date Key", "Clave a la dimension Fecha del esfuerzo",
            "int", key="FK", fk="DimDate.date_key", nullable="N", example="20260716",
            src_sys=DER, src_table="time_entries", src_field="spent_on", src_dtype="date",
            rules="CONVERT(time_entries.spent_on, formato YYYYMMDD)",
            comments="Preferir spent_on sobre tyear/tmonth/tweek, que Redmine mantiene desnormalizados (§4.3)"),
        row("hours", "Hours", "Cantidad de horas reales registradas",
            "decimal", prec="9,2", nullable="N", example="6.00, 3.50",
            src_sys=SRC, src_schema=SCH, src_table="time_entries", src_field="hours", src_dtype="float",
            rules="Copia directa con conversion a decimal(9,2)",
            comments="Unica medida aditiva de la tabla"),
        row("entry_comments", "Entry Comments", "Nota libre asociada al registro de tiempo",
            "varchar", size="1024", nullable="Y", example="'Seed time for Setup CI pipeline'",
            src_sys=SRC, src_schema=SCH, src_table="time_entries", src_field="comments", src_dtype="varchar(1024)",
            rules="Copia directa",
            comments="Atributo degenerado; evaluar excluirlo si el volumen degrada el modelo"),
    ],
}

FACT_ISSUE_HISTORY = {
    "sheet": "FactIssueHistory",
    "table_name": "FactIssueHistory",
    "table_type": "Fact (transaccional / accumulating)",
    "view_name": "vw_fact_issue_history",
    "display_name": "Historial de issues",
    "description": ("FactIssueHistory registra cada cambio auditable sobre un issue. "
                    "Grano: una fila por journal_details.id con property = 'attr' y prop_key relevante."),
    "rows": [
        row("issue_history_key", "Issue History Key", "Identificador del detalle de cambio",
            "int", key="PK", nullable="N", example="1, 2, 3",
            src_sys=SRC, src_schema=SCH, src_table="journal_details", src_field="id", src_dtype="int(11)",
            rules="Copia directa"),
        row("journal_key", "Journal Key", "Identificador de la entrada de auditoria que agrupa los cambios",
            "int", nullable="N", example="1, 2",
            src_sys=SRC, src_schema=SCH, src_table="journals", src_field="id", src_dtype="int(11)",
            rules="Copia directa",
            comments="Un journal puede contener varios journal_details del mismo evento"),
        row("issue_key", "Issue Key", "Clave al hecho Issue modificado",
            "int", key="FK", fk="FactIssue.issue_key", nullable="N", example="1, 4",
            src_sys=SRC, src_schema=SCH, src_table="journals", src_field="journalized_id", src_dtype="int(11)",
            rules="Copia directa filtrando journals.journalized_type = 'Issue'",
            comments="journals tambien audita otras entidades; el filtro es obligatorio (§4.7)"),
        row("user_key", "User Key", "Clave a la dimension Usuario que realizo el cambio",
            "int", key="FK", fk="DimUser.user_key", nullable="N", example="1, 5",
            src_sys=SRC, src_schema=SCH, src_table="journals", src_field="user_id", src_dtype="int(11)",
            rules="Copia directa"),
        row("change_date_key", "Change Date Key", "Clave a la dimension Fecha del cambio",
            "int", key="FK", fk="DimDate.date_key", nullable="N", example="20260716",
            src_sys=DER, src_table="journals", src_field="created_on", src_dtype="datetime",
            rules="CONVERT(DATE(journals.created_on), formato YYYYMMDD)"),
        row("change_timestamp", "Change Timestamp", "Marca temporal exacta del cambio",
            "datetime", nullable="N", example="2026-07-16 14:32:00",
            src_sys=SRC, src_schema=SCH, src_table="journals", src_field="created_on", src_dtype="datetime",
            rules="Copia directa",
            comments="Necesaria para ordenar cambios dentro del mismo dia al reconstruir el burndown"),
        row("change_property", "Change Property", "Atributo del issue que fue modificado",
            "varchar", size="30", nullable="N", example="'status_id', 'done_ratio'",
            src_sys=SRC, src_schema=SCH, src_table="journal_details", src_field="prop_key", src_dtype="varchar(30)",
            rules=("Copia directa filtrando journal_details.property = 'attr' AND prop_key IN "
                   "('status_id', 'done_ratio', 'fixed_version_id')"),
            comments="Restringir a las tres propiedades relevantes evita inflar la tabla (§6.2)"),
        row("old_status_key", "Old Status Key", "Estado previo al cambio",
            "int", key="FK", fk="DimIssueStatus.status_key", nullable="Y", default="-1",
            example="1, 2, -1",
            src_sys=DER, src_table="journal_details", src_field="old_value", src_dtype="longtext",
            rules="CAST(old_value AS int) cuando prop_key = 'status_id', en caso contrario NULL",
            comments="old_value es longtext; el casteo puede fallar con valores vacios"),
        row("new_status_key", "New Status Key", "Estado posterior al cambio",
            "int", key="FK", fk="DimIssueStatus.status_key", nullable="Y", default="-1",
            example="2, 5, -1",
            src_sys=DER, src_table="journal_details", src_field="value", src_dtype="longtext",
            rules="CAST(value AS int) cuando prop_key = 'status_id', en caso contrario NULL"),
        row("old_done_ratio", "Old Done Ratio", "Porcentaje de avance previo al cambio",
            "int", nullable="Y", example="0, 40",
            src_sys=DER, src_table="journal_details", src_field="old_value", src_dtype="longtext",
            rules="CAST(old_value AS int) cuando prop_key = 'done_ratio', en caso contrario NULL"),
        row("new_done_ratio", "New Done Ratio", "Porcentaje de avance posterior al cambio",
            "int", nullable="Y", example="40, 100",
            src_sys=DER, src_table="journal_details", src_field="value", src_dtype="longtext",
            rules="CAST(value AS int) cuando prop_key = 'done_ratio', en caso contrario NULL"),
        row("old_version_key", "Old Version Key", "Version asignada antes del cambio",
            "int", key="FK", fk="DimVersion.version_key", nullable="Y", default="-1",
            example="1, -1",
            src_sys=DER, src_table="journal_details", src_field="old_value", src_dtype="longtext",
            rules="CAST(old_value AS int) cuando prop_key = 'fixed_version_id', en caso contrario NULL"),
        row("new_version_key", "New Version Key", "Version asignada despues del cambio",
            "int", key="FK", fk="DimVersion.version_key", nullable="Y", default="-1",
            example="2, -1",
            src_sys=DER, src_table="journal_details", src_field="value", src_dtype="longtext",
            rules="CAST(value AS int) cuando prop_key = 'fixed_version_id', en caso contrario NULL"),
        row("is_closing_event", "Is Closing Event", "Marca el cambio que llevo el issue a estado cerrado",
            "bit", nullable="N", default="0", example="0, 1",
            src_sys=DER,
            rules=("1 cuando el estado nuevo tiene is_closed = 1 y el estado anterior tiene is_closed = 0; "
                   "requiere JOIN a issue_statuses por old_status_key y new_status_key"),
            comments="Fuente alternativa del KPI Throughput cuando closed_on resulta poco confiable (§6.3)"),
    ],
}

FACT_BURNDOWN_DAILY = {
    "sheet": "FactBurndownDaily",
    "table_name": "FactBurndownDaily",
    "table_type": "Fact (periodic snapshot)",
    "view_name": "vw_fact_burndown_daily",
    "display_name": "Burndown diario",
    "description": ("FactBurndownDaily materializa el trabajo pendiente por sprint y por dia. "
                    "Grano: una fila por version por dia de calendario dentro del rango del sprint."),
    "rows": [
        row("version_key", "Version Key", "Clave a la dimension Version sobre la que se mide el burndown",
            "int", key="PK, FK", fk="DimVersion.version_key", nullable="N", example="1, 2",
            src_sys=DER, src_table="versions", src_field="id", src_dtype="int(11)",
            rules="Producto cartesiano entre versions y los dias del calendario del sprint"),
        row("snapshot_date_key", "Snapshot Date Key", "Clave a la dimension Fecha del corte diario",
            "int", key="PK, FK", fk="DimDate.date_key", nullable="N", example="20260716",
            src_sys=DER, src_table="DimDate", src_field="date_key",
            rules="Un registro por cada dia entre el inicio del sprint y versions.effective_date"),
        row("project_key", "Project Key", "Clave a la dimension Proyecto",
            "int", key="FK", fk="DimProject.project_key", nullable="N", example="1",
            src_sys=DER, src_table="issues", src_field="project_id", src_dtype="int(11)",
            rules="Atribuir por issues.project_id de los issues de la version",
            comments="Las versiones compartidas pueden abarcar varios proyectos; no usar versions.project_id (§5.3)"),
        row("remaining_hours", "Remaining Hours", "Horas pendientes al cierre del dia",
            "decimal", prec="9,2", nullable="Y", example="24.40, 0.00",
            src_sys=DER, src_table="issues, journal_details",
            rules=("SUM(estimated_hours * (1 - done_ratio / 100)) sobre los issues abiertos, "
                   "reconstruyendo estado y done_ratio al cierre de cada dia desde journal_details"),
            comments="Escenario B del §6.2. Redmine no materializa snapshots diarios: la reconstruccion es costosa"),
        row("remaining_issues", "Remaining Issues", "Cantidad de issues abiertos al cierre del dia",
            "int", nullable="N", default="0", example="5, 2",
            src_sys=DER, src_table="issues, journal_details",
            rules="COUNT de issues con is_closed = 0 al cierre del dia, filtrando parent_id IS NULL",
            comments="Documentar la convencion elegida: solo issues raiz o solo hojas, nunca ambos (§5.2)"),
        row("completed_hours", "Completed Hours", "Horas estimadas de los issues cerrados hasta la fecha",
            "decimal", prec="9,2", nullable="N", default="0", example="0.00, 24.00",
            src_sys=DER, src_table="issues, journal_details",
            rules="SUM(estimated_hours) de los issues de la version cerrados en o antes de la fecha del corte"),
        row("ideal_remaining_hours", "Ideal Remaining Hours", "Linea ideal de referencia del burndown",
            "decimal", prec="9,2", nullable="Y", example="18.00, 12.00",
            src_sys=DER,
            rules=("Interpolacion lineal desde el total de horas estimadas al inicio del sprint "
                   "hasta 0 en versions.effective_date"),
            comments="Serie de comparacion visual; no es un dato de origen"),
        row("total_estimated_hours", "Total Estimated Hours", "Alcance total comprometido en el sprint",
            "decimal", prec="9,2", nullable="Y", example="24.00",
            src_sys=DER, src_table="issues", src_field="estimated_hours", src_dtype="float",
            rules="SUM(estimated_hours) de todos los issues de la version vigentes en la fecha del corte",
            comments="Su variacion entre dias evidencia scope creep dentro del sprint"),
    ],
}

# =============================================================================
# DIMENSIONES
# =============================================================================

DIM_DATE = {
    "sheet": "DimDate",
    "table_name": "DimDate",
    "table_type": "Dimension (generada)",
    "view_name": "DimDate",
    "display_name": "Fecha",
    "description": ("Dimension de calendario generada en Power BI, no proviene de Redmine. "
                    "Grano: una fila por dia. Rango: desde MIN(issues.created_on) hasta hoy."),
    "rows": [
        row("date_key", "Date Key", "Clave entera de la fecha en formato YYYYMMDD",
            "int", key="PK", nullable="N", example="20260725",
            src_sys=DER, rules="VALUE(FORMAT([Date], \"YYYYMMDD\"))",
            comments="Incluir un miembro -1 'Sin fecha' para las claves de hecho nulas"),
        row("date", "Date", "Fecha del calendario", "date", nullable="N", example="2026-07-25",
            src_sys=DER, rules="CALENDAR(DATE(2020,1,1), TODAY()) (§8.2)"),
        row("year", "Year", "Anio calendario", "int", nullable="N", example="2026",
            src_sys=DER, rules="YEAR([Date])"),
        row("quarter", "Quarter", "Trimestre calendario", "tinyint", nullable="N", example="3",
            src_sys=DER, rules="QUARTER([Date])"),
        row("month_num", "Month Num", "Numero de mes, de 1 a 12", "tinyint", nullable="N", example="7",
            src_sys=DER, rules="MONTH([Date])"),
        row("month_name", "Month Name", "Nombre del mes", "varchar", size="20", nullable="N",
            example="'julio'", src_sys=DER, rules="FORMAT([Date], \"MMMM\")",
            comments="Ordenar por month_num con 'Ordenar por columna' en Power BI"),
        row("year_month", "Year Month", "Anio y mes en formato ordenable", "varchar", size="7",
            nullable="N", example="'2026-07'", src_sys=DER, rules="FORMAT([Date], \"YYYY-MM\")"),
        row("iso_week", "ISO Week", "Numero de semana segun norma ISO 8601", "tinyint", nullable="N",
            example="30", src_sys=DER, rules="WEEKNUM([Date], 21)",
            comments="Eje temporal del KPI Throughput; el modo 21 corresponde a ISO (§6.3)"),
        row("iso_year_week", "ISO Year Week", "Anio y semana ISO concatenados", "varchar", size="8",
            nullable="N", example="'2026-W30'", src_sys=DER,
            rules="FORMAT(YEAR([Date]), \"0000\") & \"-W\" & FORMAT(WEEKNUM([Date],21), \"00\")",
            comments="Equivale a YEARWEEK(fecha, 3) en MySQL"),
        row("day_of_month", "Day Of Month", "Dia del mes", "tinyint", nullable="N", example="25",
            src_sys=DER, rules="DAY([Date])"),
        row("day_of_week", "Day Of Week", "Dia de la semana, 1 = lunes", "tinyint", nullable="N",
            example="6", src_sys=DER, rules="WEEKDAY([Date], 2)"),
        row("day_name", "Day Name", "Nombre del dia", "varchar", size="20", nullable="N",
            example="'sabado'", src_sys=DER, rules="FORMAT([Date], \"dddd\")"),
        row("is_weekend", "Is Weekend", "Indica si la fecha cae en fin de semana", "bit", nullable="N",
            default="0", example="0, 1", src_sys=DER, rules="IF(WEEKDAY([Date],2) > 5, 1, 0)",
            comments="Permite excluir fines de semana de la linea ideal del burndown"),
        row("is_working_day", "Is Working Day", "Indica si la fecha es dia habil", "bit", nullable="N",
            default="1", example="0, 1", src_sys=DER, rules="IF(is_weekend = 1, 0, 1)",
            comments="No contempla feriados; ampliar con un calendario laboral si el negocio lo exige"),
    ],
}

DIM_PROJECT = {
    "sheet": "DimProject",
    "table_name": "DimProject",
    "table_type": "Dimension",
    "view_name": "vw_dim_project",
    "display_name": "Proyecto",
    "description": ("Portafolio de proyectos con su jerarquia. Grano: una fila por projects.id. "
                    "Clave natural, sin historizacion (SCD tipo 1)."),
    "rows": [
        row("project_key", "Project Key", "Identificador del proyecto",
            "int", key="PK", nullable="N", example="1, 2",
            src_sys=SRC, src_schema=SCH, src_table="projects", src_field="id", src_dtype="int(11)",
            rules="Copia directa (clave natural)"),
        row("project_name", "Project Name", "Nombre visible del proyecto",
            "varchar", size="255", nullable="N", example="'BI Demo Portfolio'",
            src_sys=SRC, src_schema=SCH, src_table="projects", src_field="name", src_dtype="varchar(255)",
            rules="Copia directa"),
        row("project_identifier", "Project Identifier", "Identificador corto usado en la URL",
            "varchar", size="255", nullable="Y", example="'bi-demo'",
            src_sys=SRC, src_schema=SCH, src_table="projects", src_field="identifier", src_dtype="varchar(255)",
            rules="Copia directa"),
        row("project_description", "Project Description", "Descripcion libre del proyecto",
            "varchar", size="1000", nullable="Y", example="'Proyecto seed para validar...'",
            src_sys=SRC, src_schema=SCH, src_table="projects", src_field="description", src_dtype="text",
            rules="Truncar a 1000 caracteres"),
        row("parent_project_key", "Parent Project Key", "Proyecto padre en la jerarquia de portafolio",
            "int", nullable="Y", default="-1", example="1, -1",
            src_sys=SRC, src_schema=SCH, src_table="projects", src_field="parent_id", src_dtype="int(11)",
            rules="Copia directa. NULL se sustituye por -1",
            comments="Jerarquia padre-hijo; en Power BI aplanar con PATH() para navegacion drill-down"),
        row("project_status_code", "Project Status Code", "Codigo de estado del proyecto",
            "int", nullable="N", example="1, 5, 9",
            src_sys=SRC, src_schema=SCH, src_table="projects", src_field="status", src_dtype="int(11)",
            rules="Copia directa"),
        row("project_status_name", "Project Status Name", "Descripcion del estado del proyecto",
            "varchar", size="20", nullable="N", example="'Activo', 'Archivado'",
            src_sys=DER, src_table="projects", src_field="status", src_dtype="int(11)",
            rules="CASE status WHEN 1 THEN 'Activo' WHEN 5 THEN 'Archivado' WHEN 9 THEN 'Cerrado' END",
            comments="Codigos documentados en §4.4"),
        row("is_active", "Is Active", "Indica si el proyecto esta activo",
            "bit", nullable="N", default="1", example="0, 1",
            src_sys=DER, src_table="projects", src_field="status", src_dtype="int(11)",
            rules="CASE WHEN status = 1 THEN 1 ELSE 0 END",
            comments="Excluir los proyectos archivados de los slicers, conservandolos en los hechos historicos (§3.2)"),
        row("is_public", "Is Public", "Indica si el proyecto es de acceso publico",
            "bit", nullable="N", default="0", example="0, 1",
            src_sys=SRC, src_schema=SCH, src_table="projects", src_field="is_public", src_dtype="tinyint(1)",
            rules="Copia directa"),
        row("default_version_key", "Default Version Key", "Version asignada por defecto a los issues nuevos",
            "int", nullable="Y", default="-1", example="2, -1",
            src_sys=SRC, src_schema=SCH, src_table="projects", src_field="default_version_id", src_dtype="int(11)",
            rules="Copia directa. NULL se sustituye por -1"),
        row("created_on", "Created On", "Fecha de alta del proyecto",
            "datetime", nullable="Y", example="2026-07-25 19:10:00",
            src_sys=SRC, src_schema=SCH, src_table="projects", src_field="created_on", src_dtype="timestamp",
            rules="Copia directa", comments="Almacenado en UTC del servidor (§1.3)"),
    ],
}

DIM_USER = {
    "sheet": "DimUser",
    "table_name": "DimUser",
    "table_type": "Dimension (role-playing)",
    "view_name": "vw_dim_user",
    "display_name": "Usuario",
    "description": ("Personas que participan en el portafolio. Grano: una fila por users.id. "
                    "Actua como role-playing dimension: autor y responsable del issue."),
    "rows": [
        row("user_key", "User Key", "Identificador del usuario",
            "int", key="PK", nullable="N", example="1, 5, 6",
            src_sys=SRC, src_schema=SCH, src_table="users", src_field="id", src_dtype="int(11)",
            rules="Copia directa (clave natural)",
            comments="Agregar un miembro -1 'Sin asignar' para los issues sin responsable"),
        row("login", "Login", "Nombre de usuario para el inicio de sesion",
            "varchar", size="255", nullable="N", example="'alice', 'admin'",
            src_sys=SRC, src_schema=SCH, src_table="users", src_field="login", src_dtype="varchar(255)",
            rules="Copia directa"),
        row("first_name", "First Name", "Nombre de pila",
            "varchar", size="30", nullable="N", example="'Alice'",
            src_sys=SRC, src_schema=SCH, src_table="users", src_field="firstname", src_dtype="varchar(30)",
            rules="Copia directa"),
        row("last_name", "Last Name", "Apellido",
            "varchar", size="255", nullable="N", example="'Developer'",
            src_sys=SRC, src_schema=SCH, src_table="users", src_field="lastname", src_dtype="varchar(255)",
            rules="Copia directa"),
        row("full_name", "Full Name", "Nombre completo para mostrar en los visuales",
            "varchar", size="290", nullable="N", example="'Alice Developer'",
            src_sys=DER, src_table="users", src_field="firstname, lastname",
            rules="CONCAT(firstname, ' ', lastname)",
            comments="Etiqueta preferida en slicers y ejes de los visuales"),
        row("email", "Email", "Direccion de correo principal",
            "varchar", size="255", nullable="Y", example="'alice@example.com'",
            src_sys=SRC, src_schema=SCH, src_table="email_addresses", src_field="address", src_dtype="varchar(255)",
            rules="JOIN email_addresses ON user_id = users.id WHERE is_default = 1",
            comments="Desde Redmine 4 el correo vive en email_addresses, no en users"),
        row("user_type", "User Type", "Tipo de registro: usuario, grupo o anonimo",
            "varchar", size="255", nullable="Y", example="'User', 'Group'",
            src_sys=SRC, src_schema=SCH, src_table="users", src_field="type", src_dtype="varchar(255)",
            rules="Copia directa",
            comments="Filtrar type = 'User' para excluir grupos y cuentas de sistema de los analisis por persona"),
        row("is_admin", "Is Admin", "Indica si el usuario tiene privilegios de administrador",
            "bit", nullable="N", default="0", example="0, 1",
            src_sys=SRC, src_schema=SCH, src_table="users", src_field="admin", src_dtype="tinyint(1)",
            rules="Copia directa"),
        row("status_code", "Status Code", "Codigo de estado de la cuenta",
            "int", nullable="N", example="1, 3",
            src_sys=SRC, src_schema=SCH, src_table="users", src_field="status", src_dtype="int(11)",
            rules="Copia directa"),
        row("is_active", "Is Active", "Indica si la cuenta esta activa",
            "bit", nullable="N", default="1", example="0, 1",
            src_sys=DER, src_table="users", src_field="status", src_dtype="int(11)",
            rules="CASE WHEN status = 1 THEN 1 ELSE 0 END",
            comments="Excluir inactivos de los slicers y conservarlos en los hechos historicos (§5.4)"),
        row("last_login_on", "Last Login On", "Fecha del ultimo inicio de sesion",
            "datetime", nullable="Y", example="2026-07-25 18:00:00",
            src_sys=SRC, src_schema=SCH, src_table="users", src_field="last_login_on", src_dtype="datetime",
            rules="Copia directa"),
        row("created_on", "Created On", "Fecha de alta de la cuenta",
            "datetime", nullable="Y", example="2026-07-25 19:14:00",
            src_sys=SRC, src_schema=SCH, src_table="users", src_field="created_on", src_dtype="timestamp",
            rules="Copia directa"),
        row("[EXCLUIDO]", "N/A", "Campos de autenticacion que no deben importarse",
            "N/A", nullable="N/A", example="N/A",
            src_sys=SRC, src_schema=SCH, src_table="users",
            src_field="hashed_password, salt, twofa_totp_key", src_dtype="varchar",
            rules="NO EXTRAER",
            comments="Regla de seguridad explicita: no incluir credenciales en el modelo semantico (§3.2, §4.8)"),
    ],
}

DIM_ISSUE_STATUS = {
    "sheet": "DimIssueStatus",
    "table_name": "DimIssueStatus",
    "table_type": "Dimension",
    "view_name": "vw_dim_issue_status",
    "display_name": "Estado del issue",
    "description": ("Catalogo de estados del flujo de trabajo. Grano: una fila por issue_statuses.id. "
                    "Define la regla canonica de tarea terminada."),
    "rows": [
        row("status_key", "Status Key", "Identificador del estado",
            "int", key="PK", nullable="N", example="1, 2, 5",
            src_sys=SRC, src_schema=SCH, src_table="issue_statuses", src_field="id", src_dtype="int(11)",
            rules="Copia directa (clave natural)"),
        row("status_name", "Status Name", "Etiqueta del estado",
            "varchar", size="30", nullable="N", example="'New', 'In Progress', 'Closed'",
            src_sys=SRC, src_schema=SCH, src_table="issue_statuses", src_field="name", src_dtype="varchar(30)",
            rules="Copia directa",
            comments="Valores cargados en el entorno: New, In Progress, Resolved, Feedback, Closed, Rejected"),
        row("status_description", "Status Description", "Descripcion del estado",
            "varchar", size="255", nullable="Y", example="NULL",
            src_sys=SRC, src_schema=SCH, src_table="issue_statuses", src_field="description",
            src_dtype="varchar(255)", rules="Copia directa"),
        row("is_closed", "Is Closed", "Indica si el estado se considera terminado",
            "bit", nullable="N", default="0", example="0, 1",
            src_sys=SRC, src_schema=SCH, src_table="issue_statuses", src_field="is_closed",
            src_dtype="tinyint(1)", rules="Copia directa",
            comments="Regla de negocio central de Throughput y Velocity. Closed y Rejected tienen is_closed = 1 (§5.1)"),
        row("status_position", "Status Position", "Orden del estado en el tablero",
            "int", nullable="Y", example="1, 2, 5",
            src_sys=SRC, src_schema=SCH, src_table="issue_statuses", src_field="position", src_dtype="int(11)",
            rules="Copia directa",
            comments="Usar como columna de ordenamiento de status_name en Power BI"),
        row("default_done_ratio", "Default Done Ratio", "Avance por defecto al entrar en el estado",
            "int", nullable="Y", example="NULL, 100",
            src_sys=SRC, src_schema=SCH, src_table="issue_statuses", src_field="default_done_ratio",
            src_dtype="int(11)", rules="Copia directa"),
    ],
}

DIM_VERSION = {
    "sheet": "DimVersion",
    "table_name": "DimVersion",
    "table_type": "Dimension",
    "view_name": "vw_dim_version",
    "display_name": "Version / Sprint",
    "description": ("Milestones nativos de Redmine que operan como sprint o release. "
                    "Grano: una fila por versions.id. Es la dimension de sprint por defecto (§7.1)."),
    "rows": [
        row("version_key", "Version Key", "Identificador de la version",
            "int", key="PK", nullable="N", example="1, 2",
            src_sys=SRC, src_schema=SCH, src_table="versions", src_field="id", src_dtype="int(11)",
            rules="Copia directa (clave natural)",
            comments="Agregar un miembro -1 'Sin version' para los issues sin sprint asignado"),
        row("project_key", "Project Key", "Proyecto propietario de la version",
            "int", key="FK", fk="DimProject.project_key", nullable="N", example="1",
            src_sys=SRC, src_schema=SCH, src_table="versions", src_field="project_id", src_dtype="int(11)",
            rules="Copia directa",
            comments="Proyecto duenio de la version; para atribuir hechos usar issues.project_id (§5.3)"),
        row("version_name", "Version Name", "Nombre del sprint o release",
            "varchar", size="255", nullable="N", example="'Sprint 1', 'Sprint 2'",
            src_sys=SRC, src_schema=SCH, src_table="versions", src_field="name", src_dtype="varchar(255)",
            rules="Copia directa"),
        row("version_description", "Version Description", "Descripcion del sprint o release",
            "varchar", size="255", nullable="Y", example="'Sprint en curso'",
            src_sys=SRC, src_schema=SCH, src_table="versions", src_field="description",
            src_dtype="varchar(255)", rules="Copia directa"),
        row("effective_date", "Effective Date", "Fecha objetivo del milestone",
            "date", nullable="Y", example="2026-08-02",
            src_sys=SRC, src_schema=SCH, src_table="versions", src_field="effective_date", src_dtype="date",
            rules="Copia directa",
            comments=("Unica fecha disponible de forma nativa: no hay fecha de inicio de sprint. "
                      "Es la referencia del Schedule Variance a nivel version (§6.1)")),
        row("version_status", "Version Status", "Estado del ciclo de vida de la version",
            "varchar", size="20", nullable="Y", example="'open', 'closed'",
            src_sys=SRC, src_schema=SCH, src_table="versions", src_field="status", src_dtype="varchar(255)",
            rules="Copia directa",
            comments="Valores posibles: open, locked, closed"),
        row("is_open", "Is Open", "Indica si la version admite nuevos issues",
            "bit", nullable="N", default="1", example="0, 1",
            src_sys=DER, src_table="versions", src_field="status", src_dtype="varchar(255)",
            rules="CASE WHEN status = 'open' THEN 1 ELSE 0 END",
            comments="Redmine impide asignar issues a versiones cerradas: relevante para la carga"),
        row("sharing", "Sharing", "Alcance de comparticion de la version entre proyectos",
            "varchar", size="20", nullable="N", default="'none'", example="'none', 'tree'",
            src_sys=SRC, src_schema=SCH, src_table="versions", src_field="sharing", src_dtype="varchar(255)",
            rules="Copia directa",
            comments="Valores: none, descendants, hierarchy, tree, system. Distinto de none implica riesgo de atribucion cruzada (§5.3)"),
        row("created_on", "Created On", "Fecha de creacion de la version",
            "datetime", nullable="Y", example="2026-07-25 19:15:00",
            src_sys=SRC, src_schema=SCH, src_table="versions", src_field="created_on", src_dtype="timestamp",
            rules="Copia directa"),
    ],
}

DIM_TRACKER = {
    "sheet": "DimTracker",
    "table_name": "DimTracker",
    "table_type": "Dimension",
    "view_name": "vw_dim_tracker",
    "display_name": "Tracker",
    "description": "Tipo de issue segun el flujo de trabajo. Grano: una fila por trackers.id.",
    "rows": [
        row("tracker_key", "Tracker Key", "Identificador del tracker",
            "int", key="PK", nullable="N", example="1, 2, 3",
            src_sys=SRC, src_schema=SCH, src_table="trackers", src_field="id", src_dtype="int(11)",
            rules="Copia directa (clave natural)"),
        row("tracker_name", "Tracker Name", "Nombre del tipo de issue",
            "varchar", size="30", nullable="N", example="'Bug', 'Feature', 'Support'",
            src_sys=SRC, src_schema=SCH, src_table="trackers", src_field="name", src_dtype="varchar(30)",
            rules="Copia directa",
            comments="Valores cargados en el entorno: Bug, Feature, Support"),
        row("tracker_description", "Tracker Description", "Descripcion del tracker",
            "varchar", size="255", nullable="Y", example="NULL",
            src_sys=SRC, src_schema=SCH, src_table="trackers", src_field="description",
            src_dtype="varchar(255)", rules="Copia directa"),
        row("is_in_roadmap", "Is In Roadmap", "Indica si el tracker aparece en la hoja de ruta",
            "bit", nullable="N", default="1", example="0, 1",
            src_sys=SRC, src_schema=SCH, src_table="trackers", src_field="is_in_roadmap",
            src_dtype="tinyint(1)", rules="Copia directa"),
        row("tracker_position", "Tracker Position", "Orden de presentacion del tracker",
            "int", nullable="Y", example="1, 2, 3",
            src_sys=SRC, src_schema=SCH, src_table="trackers", src_field="position", src_dtype="int(11)",
            rules="Copia directa"),
        row("default_status_key", "Default Status Key", "Estado inicial de los issues de este tracker",
            "int", key="FK", fk="DimIssueStatus.status_key", nullable="Y", default="-1", example="1",
            src_sys=SRC, src_schema=SCH, src_table="trackers", src_field="default_status_id",
            src_dtype="int(11)", rules="Copia directa"),
    ],
}

DIM_PRIORITY = {
    "sheet": "DimPriority",
    "table_name": "DimPriority",
    "table_type": "Dimension",
    "view_name": "vw_dim_priority",
    "display_name": "Prioridad",
    "description": ("Prioridad operativa nativa de Redmine. Grano: una fila por enumerations.id "
                    "filtrando type = 'IssuePriority'. Funciona como proxy de severidad (§7.2)."),
    "rows": [
        row("priority_key", "Priority Key", "Identificador de la prioridad",
            "int", key="PK", nullable="N", example="1, 2, 3",
            src_sys=SRC, src_schema=SCH, src_table="enumerations", src_field="id", src_dtype="int(11)",
            rules="Copia directa filtrando type = 'IssuePriority'",
            comments="enumerations es una tabla polimorfica: sin el filtro se mezclan actividades y categorias"),
        row("priority_name", "Priority Name", "Etiqueta de la prioridad",
            "varchar", size="30", nullable="N", example="'Low', 'Normal', 'High'",
            src_sys=SRC, src_schema=SCH, src_table="enumerations", src_field="name", src_dtype="varchar(30)",
            rules="Copia directa",
            comments="Valores cargados: Low, Normal, High, Urgent, Immediate"),
        row("priority_position", "Priority Position", "Orden jerarquico de la prioridad",
            "int", nullable="Y", example="1, 2, 3",
            src_sys=SRC, src_schema=SCH, src_table="enumerations", src_field="position", src_dtype="int(11)",
            rules="Copia directa",
            comments="Ordenar priority_name por esta columna para que los visuales respeten la escala"),
        row("is_default", "Is Default", "Indica si es la prioridad asignada por defecto",
            "bit", nullable="N", default="0", example="0, 1",
            src_sys=SRC, src_schema=SCH, src_table="enumerations", src_field="is_default",
            src_dtype="tinyint(1)", rules="Copia directa"),
        row("is_active", "Is Active", "Indica si la prioridad esta habilitada",
            "bit", nullable="N", default="1", example="0, 1",
            src_sys=SRC, src_schema=SCH, src_table="enumerations", src_field="active",
            src_dtype="tinyint(1)", rules="Copia directa"),
    ],
}

DIM_SEVERITY = {
    "sheet": "DimSeverity",
    "table_name": "DimSeverity",
    "table_type": "Dimension (CONDICIONAL)",
    "view_name": "vw_dim_severity",
    "display_name": "Severidad",
    "description": ("Severidad de negocio implementada como campo personalizado (patron EAV). "
                    "CONDICIONAL: verificado el 2026-07-25, el campo 'Severidad' NO existe en el entorno. "
                    "Mientras tanto el dashboard usa DimPriority como proxy (§7.2)."),
    "rows": [
        row("severity_key", "Severity Key", "Identificador del valor de severidad",
            "int", key="PK", nullable="N", example="1, 2, 3",
            src_sys=SRC, src_schema=SCH, src_table="custom_field_enumerations", src_field="id",
            src_dtype="int(11)",
            rules=("Copia directa cuando el campo se define con formato 'key/value list'. "
                   "Si se define como lista simple, generar la clave con RANKX sobre los valores distintos"),
            comments="Agregar un miembro -1 'Sin severidad' para los issues sin valor cargado"),
        row("custom_field_key", "Custom Field Key", "Campo personalizado al que pertenece el valor",
            "int", key="FK", nullable="N", example="1",
            src_sys=SRC, src_schema=SCH, src_table="custom_field_enumerations", src_field="custom_field_id",
            src_dtype="int(11)",
            rules="Copia directa filtrando custom_fields.name = 'Severidad' AND custom_fields.type = 'IssueCustomField'"),
        row("severity_name", "Severity Name", "Etiqueta de la severidad",
            "varchar", size="255", nullable="N", example="'Critical', 'Major', 'Minor'",
            src_sys=SRC, src_schema=SCH, src_table="custom_field_enumerations", src_field="name",
            src_dtype="varchar(255)", rules="Copia directa"),
        row("severity_position", "Severity Position", "Orden de la escala de severidad",
            "int", nullable="N", example="1, 2, 3",
            src_sys=SRC, src_schema=SCH, src_table="custom_field_enumerations", src_field="position",
            src_dtype="int(11)", rules="Copia directa",
            comments="Columna de ordenamiento de severity_name"),
        row("is_active", "Is Active", "Indica si el valor de severidad esta habilitado",
            "bit", nullable="N", default="1", example="0, 1",
            src_sys=SRC, src_schema=SCH, src_table="custom_field_enumerations", src_field="active",
            src_dtype="tinyint(1)", rules="Copia directa"),
        row("[VALIDACION]", "N/A", "Control previo a la carga de esta dimension",
            "N/A", nullable="N/A", example="N/A",
            src_sys=SRC, src_schema=SCH, src_table="custom_fields", src_field="name", src_dtype="varchar(30)",
            rules="SELECT * FROM custom_fields WHERE name = 'Severidad'",
            comments="Si no devuelve filas, omitir la dimension y usar DimPriority con nota al pie (§9.1)"),
    ],
}

DIM_MEMBER = {
    "sheet": "DimMember",
    "table_name": "DimMember",
    "table_type": "Dimension puente (bridge)",
    "view_name": "vw_dim_member",
    "display_name": "Membresia",
    "description": ("Tabla puente que resuelve la relacion muchos a muchos entre usuarios, proyectos y roles. "
                    "Grano: una fila por member_roles.id, es decir por usuario, proyecto y rol."),
    "rows": [
        row("member_role_key", "Member Role Key", "Identificador de la asignacion usuario-proyecto-rol",
            "int", key="PK", nullable="N", example="1, 2, 3",
            src_sys=SRC, src_schema=SCH, src_table="member_roles", src_field="id", src_dtype="int(11)",
            rules="Copia directa",
            comments="El grano es member_roles y no members: un usuario puede tener varios roles en un proyecto"),
        row("member_key", "Member Key", "Identificador de la membresia usuario-proyecto",
            "int", nullable="N", example="1, 2",
            src_sys=SRC, src_schema=SCH, src_table="member_roles", src_field="member_id", src_dtype="int(11)",
            rules="Copia directa"),
        row("user_key", "User Key", "Clave a la dimension Usuario",
            "int", key="FK", fk="DimUser.user_key", nullable="N", example="5, 6, 7",
            src_sys=SRC, src_schema=SCH, src_table="members", src_field="user_id", src_dtype="int(11)",
            rules="JOIN members ON members.id = member_roles.member_id"),
        row("project_key", "Project Key", "Clave a la dimension Proyecto",
            "int", key="FK", fk="DimProject.project_key", nullable="N", example="1",
            src_sys=SRC, src_schema=SCH, src_table="members", src_field="project_id", src_dtype="int(11)",
            rules="JOIN members ON members.id = member_roles.member_id"),
        row("role_key", "Role Key", "Identificador del rol desempeniado",
            "int", nullable="N", example="3, 4, 5",
            src_sys=SRC, src_schema=SCH, src_table="member_roles", src_field="role_id", src_dtype="int(11)",
            rules="Copia directa"),
        row("role_name", "Role Name", "Nombre del rol",
            "varchar", size="255", nullable="N", example="'Manager', 'Developer'",
            src_sys=SRC, src_schema=SCH, src_table="roles", src_field="name", src_dtype="varchar(255)",
            rules="JOIN roles ON roles.id = member_roles.role_id",
            comments="Roles cargados: Manager, Developer, Reporter, ademas de Non member y Anonymous"),
        row("is_builtin_role", "Is Builtin Role", "Indica si el rol es interno de Redmine",
            "bit", nullable="N", default="0", example="0, 1",
            src_sys=DER, src_table="roles", src_field="builtin", src_dtype="int(11)",
            rules="CASE WHEN roles.builtin > 0 THEN 1 ELSE 0 END",
            comments="Excluir roles internos (Non member, Anonymous) de los analisis de equipo"),
        row("is_inherited", "Is Inherited", "Indica si el rol se hereda de un proyecto padre o grupo",
            "bit", nullable="N", default="0", example="0, 1",
            src_sys=DER, src_table="member_roles", src_field="inherited_from", src_dtype="int(11)",
            rules="CASE WHEN member_roles.inherited_from IS NULL THEN 0 ELSE 1 END"),
        row("member_created_on", "Member Created On", "Fecha de alta del usuario en el proyecto",
            "datetime", nullable="Y", example="2026-07-25 19:15:00",
            src_sys=SRC, src_schema=SCH, src_table="members", src_field="created_on", src_dtype="timestamp",
            rules="Copia directa",
            comments="Permite acotar Velocity per capita a los miembros activos durante el sprint (§6.5)"),
    ],
}

DIM_SPRINT = {
    "sheet": "DimSprint",
    "table_name": "DimSprint",
    "table_type": "Dimension (CONDICIONAL - plugin)",
    "view_name": "vw_dim_sprint",
    "display_name": "Sprint (Agile)",
    "description": ("Sprints del plugin redmine_agile, con fechas de inicio y fin explicitas. "
                    "CONDICIONAL: verificado el 2026-07-25 con SHOW TABLES LIKE 'agile%', el plugin NO esta "
                    "instalado. Mientras tanto se usa DimVersion como sprint nativo (§7.1)."),
    "rows": [
        row("sprint_key", "Sprint Key", "Identificador del sprint",
            "int", key="PK", nullable="N", example="1, 2",
            src_sys=SRC, src_schema=SCH, src_table="agile_sprints", src_field="id", src_dtype="int(11)",
            rules="Copia directa (clave natural)",
            comments="Nombres de tabla y campo a confirmar tras instalar el plugin"),
        row("project_key", "Project Key", "Proyecto propietario del sprint",
            "int", key="FK", fk="DimProject.project_key", nullable="N", example="1",
            src_sys=SRC, src_schema=SCH, src_table="agile_sprints", src_field="project_id", src_dtype="int(11)",
            rules="Copia directa"),
        row("sprint_name", "Sprint Name", "Nombre del sprint",
            "varchar", size="255", nullable="N", example="'Sprint 1'",
            src_sys=SRC, src_schema=SCH, src_table="agile_sprints", src_field="name", src_dtype="varchar(255)",
            rules="Copia directa"),
        row("start_date", "Start Date", "Fecha de inicio del sprint",
            "date", nullable="Y", example="2026-07-13",
            src_sys=SRC, src_schema=SCH, src_table="agile_sprints", src_field="sprint_start_date",
            src_dtype="date", rules="Copia directa",
            comments="Ventaja frente a versions, que solo expone effective_date (§7.1)"),
        row("end_date", "End Date", "Fecha de fin del sprint",
            "date", nullable="Y", example="2026-07-26",
            src_sys=SRC, src_schema=SCH, src_table="agile_sprints", src_field="sprint_end_date",
            src_dtype="date", rules="Copia directa",
            comments="Permite construir el eje del burndown sin estimar el inicio del sprint"),
        row("sprint_status", "Sprint Status", "Estado del sprint",
            "varchar", size="20", nullable="Y", example="'open', 'closed'",
            src_sys=SRC, src_schema=SCH, src_table="agile_sprints", src_field="status", src_dtype="varchar(255)",
            rules="Copia directa"),
        row("[VALIDACION]", "N/A", "Control previo a la carga de esta dimension",
            "N/A", nullable="N/A", example="N/A",
            src_sys=SRC, src_schema=SCH, src_table="information_schema", src_field="TABLES",
            rules="SHOW TABLES LIKE 'agile%'",
            comments="Si no devuelve filas, omitir la dimension y modelar el sprint con DimVersion (§9.1)"),
    ],
}

TABLES = [
    FACT_ISSUE, FACT_TIME_ENTRY, FACT_ISSUE_HISTORY, FACT_BURNDOWN_DAILY,
    DIM_DATE, DIM_PROJECT, DIM_USER, DIM_ISSUE_STATUS, DIM_VERSION,
    DIM_TRACKER, DIM_PRIORITY, DIM_SEVERITY, DIM_MEMBER, DIM_SPRINT,
]

# --- Estilos ------------------------------------------------------------------
YELLOW = PatternFill("solid", fgColor="FFF2A8")
TARGET_FILL = PatternFill("solid", fgColor="D9E1F2")
SOURCE_FILL = PatternFill("solid", fgColor="E2EFDA")
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
INDEX_FILL = PatternFill("solid", fgColor="BDD7EE")

THIN = Side(style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BOLD = Font(bold=True, size=10)
NORMAL = Font(size=9)
TITLE = Font(bold=True, size=12)

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

HEADER_ROWS = 5
GROUP_ROW = 7
COL_HEADER_ROW = 8
FIRST_DATA_ROW = 9
N_TARGET = len(TARGET_COLS)
N_TOTAL = N_TARGET + len(SOURCE_COLS)


def build_sheet(wb, spec):
    ws = wb.create_sheet(spec["sheet"])

    meta = [
        ("Table Name", spec["table_name"]),
        ("Table Type", spec["table_type"]),
        ("View Name", spec["view_name"]),
        ("Display Name", spec["display_name"]),
        ("Description", spec["description"]),
    ]
    for idx, (label, value) in enumerate(meta, start=1):
        cell_label = ws.cell(row=idx, column=1, value=label)
        cell_label.font = BOLD
        cell_label.fill = YELLOW
        cell_value = ws.cell(row=idx, column=2, value=value)
        cell_value.font = TITLE if idx == 1 else NORMAL
        cell_value.alignment = WRAP_TOP
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=min(9, N_TOTAL))
    ws.row_dimensions[5].height = 30

    ws.merge_cells(start_row=GROUP_ROW, start_column=1, end_row=GROUP_ROW, end_column=N_TARGET)
    ws.merge_cells(start_row=GROUP_ROW, start_column=N_TARGET + 1, end_row=GROUP_ROW, end_column=N_TOTAL)
    target_cell = ws.cell(row=GROUP_ROW, column=1, value="Target")
    target_cell.font = BOLD
    target_cell.fill = TARGET_FILL
    target_cell.alignment = CENTER
    source_cell = ws.cell(row=GROUP_ROW, column=N_TARGET + 1, value="Source")
    source_cell.font = BOLD
    source_cell.fill = SOURCE_FILL
    source_cell.alignment = CENTER
    for col in range(1, N_TOTAL + 1):
        ws.cell(row=GROUP_ROW, column=col).border = BORDER

    all_cols = TARGET_COLS + SOURCE_COLS
    for col_idx, (name, width) in enumerate(all_cols, start=1):
        cell = ws.cell(row=COL_HEADER_ROW, column=col_idx, value=name)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[COL_HEADER_ROW].height = 30

    for r_idx, data in enumerate(spec["rows"], start=FIRST_DATA_ROW):
        for c_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = NORMAL
            cell.alignment = WRAP_TOP
            cell.border = BORDER
        ws.row_dimensions[r_idx].height = 30

    ws.freeze_panes = ws.cell(row=FIRST_DATA_ROW, column=1)
    ws.auto_filter.ref = (f"A{COL_HEADER_ROW}:"
                          f"{get_column_letter(N_TOTAL)}{FIRST_DATA_ROW + len(spec['rows']) - 1}")
    return ws


def build_index(wb):
    ws = wb.create_sheet("Indice", 0)
    ws["A1"] = "Mapeo origen-destino (Source-to-Target Mapping)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Modelo BI de Redmine - g6-ceid-redmine"
    ws["A2"].font = Font(size=11)
    ws["A3"] = ("Origen: MySQL 5.7, base redmine_db. Tipos de dato tomados de information_schema. "
                "Estrategia de claves: naturales (SCD tipo 1). Referencias §N remiten a docs/redmine-bi-model.md.")
    ws["A3"].font = NORMAL
    ws["A3"].alignment = WRAP_TOP
    ws.merge_cells("A3:E3")
    ws.row_dimensions[3].height = 30

    headers = ["#", "Hoja", "Tipo", "Grano", "Columnas"]
    widths = [5, 24, 32, 76, 11]
    for col_idx, (name, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=5, column=col_idx, value=name)
        cell.font = BOLD
        cell.fill = INDEX_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for i, spec in enumerate(TABLES, start=1):
        r = 5 + i
        grain = spec["description"].split("Grano:")[-1].split(".")[0].strip() if "Grano:" in spec["description"] else "-"
        values = [i, spec["sheet"], spec["table_type"], grain, len(spec["rows"])]
        for c_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c_idx, value=value)
            cell.font = NORMAL
            cell.alignment = WRAP_TOP
            cell.border = BORDER
        ws.row_dimensions[r].height = 28

    note_row = 6 + len(TABLES) + 1
    ws.cell(row=note_row, column=1, value="Notas de alcance").font = BOLD
    notes = [
        "DimSeverity y DimSprint estan documentadas pero NO son cargables hoy: no existe el campo personalizado "
        "'Severidad' ni el plugin redmine_agile. Cada hoja incluye una fila [VALIDACION] con el control previo.",
        "FactTimeEntry declara activity_key hacia DimActivity, que el modelo actual no define. "
        "Fuente sugerida: enumerations WHERE type = 'TimeEntryActivity'.",
        "FactBurndownDaily es integramente derivada: no tiene tabla de origen directa y exige reconstruir "
        "el historial desde journal_details.",
        "Las claves foraneas admiten un miembro -1 para representar los valores nulos del origen.",
    ]
    for i, note in enumerate(notes):
        r = note_row + 1 + i
        cell = ws.cell(row=r, column=1, value=f"{i + 1}. {note}")
        cell.font = NORMAL
        cell.alignment = WRAP_TOP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 30
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)
    for spec in TABLES:
        build_sheet(wb, spec)
    build_index(wb)
    wb.active = 0

    out = Path(__file__).parent / "redmine-stm.xlsx"
    wb.save(out)
    total = sum(len(s["rows"]) for s in TABLES)
    print(f"OK: {out}")
    print(f"  hojas: {len(TABLES) + 1} (indice + {len(TABLES)} tablas)")
    print(f"  filas de mapeo: {total}")


if __name__ == "__main__":
    main()
